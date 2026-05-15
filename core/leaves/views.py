from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import LeaveRequest
from django.shortcuts import get_object_or_404
from users.models import Profile
from django.contrib import messages
from datetime import timedelta, datetime, date
from django.db.models import Count, Q, Sum, Prefetch
from django.utils import timezone
from leaves.models import LeaveRequest
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_protect
import json
import calendar
from django.contrib.auth.models import User
from django.utils.timezone import now
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from datetime import datetime
from .services.email_service import LeaveEmailService

@login_required
def request_leave(request):
    profile = request.user.profile   # Ensure profile exists

    if request.method == 'POST':
        start = request.POST['start_date']
        end = request.POST['end_date']
        leave_type = request.POST['leave_type']
        duration = request.POST.get('duration', 'full')
        reason = request.POST['reason']

        LeaveRequest.objects.create(
            user=request.user,
            start_date=start,
            end_date=end,
            leave_type=leave_type,
            duration=duration,
            reason=reason
        )
        messages.success(request, "Leave request submitted successfully.")
    
        return redirect('/calendar/')

    return render(request, 'leaves/request_leave.html',
        {'profile': profile})

@login_required
def team_analytics(request):
    """
    Comprehensive team analytics dashboard
    Shows team statistics, leave patterns, and trends
    Includes HALF-DAY support with duration field
    Only accessible to managers and directors
    """
    
    # Security check - Only managers and directors
    profile = request.user.profile
    if profile.role not in ['manager', 'director']:
        return redirect('employee_dashboard')
    
    today = timezone.now().date()
    current_month_start = today.replace(day=1)
    next_month = current_month_start + timedelta(days=32)
    current_month_end = next_month.replace(day=1) - timedelta(days=1)
    
    # ============ TEAM COUNTS ============
    global_count = Profile.objects.filter(team='Global').count()
    tech_count = Profile.objects.filter(team='Tech').count()
    dcm_count = Profile.objects.filter(team='DCM').count()
    total_members = tech_count + dcm_count + global_count
    
    # Calculate percentages
    tech_percentage = round((tech_count / total_members * 100)) if total_members > 0 else 0
    dcm_percentage = round((dcm_count / total_members * 100)) if total_members > 0 else 0
    
    # ============ HELPER FUNCTION - CALCULATE LEAVE DAYS (WITH HALF-DAY SUPPORT) ============
    def calculate_leave_days(leave_obj):
        """Calculate total days, accounting for half days"""
        total_days = (leave_obj.end_date - leave_obj.start_date).days + 1
        
        # If it's a half day, multiply by 0.5
        if leave_obj.duration in ['first_half', 'second_half']:
            return total_days * 0.5
        
        # Otherwise return full days
        return total_days
    
    # ============ LEAVE STATISTICS ============
    
    # Total leaves used (count of approved leave REQUESTS, not days)
    total_leaves_requests = LeaveRequest.objects.filter(
        status='approved'
    ).count()
    
    # Pending approvals
    pending_approvals = LeaveRequest.objects.filter(
        status='pending'
    ).count()
    
    # Rejected leaves
    rejected_leaves = LeaveRequest.objects.filter(
        status='rejected'
    ).count()
    
    # Average leaves per employee (only approved)
    avg_leaves_per_emp = round(
        total_leaves_requests / total_members
    ) if total_members > 0 else 0
    
    # ============ PEOPLE ON LEAVE TODAY ============
    on_leave_today = LeaveRequest.objects.filter(
        status='approved',
        start_date__lte=today,
        end_date__gte=today
    ).values('user').distinct().count()
    
    # ============ CURRENT MONTH LEAVES BY TEAM (WITH HALF-DAY CALCULATION) ============
    leaves_by_team_requests = LeaveRequest.objects.filter(
        status='approved',
        start_date__gte=current_month_start,
        start_date__lte=current_month_end
    ).values('user__profile__team').annotate(
        total=Count('id')
    )
    
    # Add days calculation for team leaves
    leaves_by_team = []
    for team_data in leaves_by_team_requests:
        team_name = team_data['user__profile__team']
        team_leaves = LeaveRequest.objects.filter(
            status='approved',
            start_date__gte=current_month_start,
            start_date__lte=current_month_end,
            user__profile__team=team_name
        )
        
        days_total = 0
        for leave in team_leaves:
            days_total += calculate_leave_days(leave)
        
        leaves_by_team.append({
            'team': team_name,
            'count': team_data['total'],
            'days': round(days_total, 1)
        })
    
    # ============ MONTHLY TREND DATA (LAST 6 MONTHS) ============
    monthly_trend = []
    monthly_trend_days = []  # NEW: Track days, not just count
    months = []
    
    for i in range(5, -1, -1):  # Last 6 months
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        
        if i == 0:
            month_end = today
        else:
            next_m = month_start + timedelta(days=32)
            month_end = next_m.replace(day=1) - timedelta(days=1)
        
        # Count leaves AND calculate days
        month_leaves = LeaveRequest.objects.filter(
            status='approved',
            start_date__gte=month_start,
            start_date__lte=month_end
        )
        
        leaves_count = month_leaves.count()
        days_count = sum(calculate_leave_days(leave) for leave in month_leaves)
        
        monthly_trend.append(leaves_count)
        monthly_trend_days.append(round(days_count, 1))
        months.append(month_start.strftime('%b'))
    
    # Get max value for scaling percentages in chart
    max_leaves_month = max(monthly_trend) if monthly_trend else 1
    max_days_month = max(monthly_trend_days) if monthly_trend_days else 1
    
    # Scale to percentage
    monthly_percentages = [
        round((leaves / max_leaves_month * 100)) if max_leaves_month > 0 else 0
        for leaves in monthly_trend
    ]
    
    monthly_days_percentages = [
        round((days / max_days_month * 100)) if max_days_month > 0 else 0
        for days in monthly_trend_days
    ]
    
    # ============ LEAVE DISTRIBUTION BY TYPE (WITH HALF-DAY CALCULATION) ============
    leave_type_data = []
    leave_types_raw = LeaveRequest.objects.filter(
        status='approved'
    ).values('leave_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    for lt in leave_types_raw:
        leaves_of_type = LeaveRequest.objects.filter(
            status='approved',
            leave_type=lt['leave_type']
        )
        days_total = sum(calculate_leave_days(leave) for leave in leaves_of_type)
        
        leave_type_data.append({
            'leave_type': lt['leave_type'],
            'count': lt['count'],
            'days': round(days_total, 1)
        })
    
    # ============ TEAM-WISE PENDING LEAVES ============
    pending_by_team = LeaveRequest.objects.filter(
        status='pending'
    ).values('user__profile__team').annotate(
        count=Count('id')
    )
    
    # ============ EMPLOYEES WITH MOST LEAVES (BY DAYS, NOT COUNT) ============
    # Get all approved leaves and group by employee
    all_approved = LeaveRequest.objects.filter(
        status='approved'
    ).select_related('user', 'user__profile')
    
    employee_days = {}
    for leave in all_approved:
        emp_key = (leave.user.id, leave.user.first_name, leave.user.last_name, leave.user.profile.team)
        if emp_key not in employee_days:
            employee_days[emp_key] = 0
        employee_days[emp_key] += calculate_leave_days(leave)
    
    # Convert to list and sort by days
    top_leave_takers = [
        {
            'user_id': k[0],
            'user__first_name': k[1],
            'user__last_name': k[2],
            'user__profile__team': k[3],
            'total_days': round(v, 1)
        }
        for k, v in sorted(employee_days.items(), key=lambda x: x[1], reverse=True)[:5]
    ]
    
    # ============ UPCOMING LEAVES (NEXT 30 DAYS) ============
    upcoming_leaves = LeaveRequest.objects.filter(
        status='approved',
        start_date__gte=today,
        start_date__lte=today + timedelta(days=30)
    ).select_related('user__profile').order_by('start_date')[:5]
    
    # ============ CALCULATE TOTAL DAYS CONSUMED (WITH HALF-DAY SUPPORT) ============
    total_days_consumed = 0
    for leave in LeaveRequest.objects.filter(status='approved'):
        days = calculate_leave_days(leave)
        total_days_consumed += days
    
    total_days_consumed = round(total_days_consumed, 1)
    
    # ============ ATTENDANCE RATE BY TEAM (WITH HALF-DAY CALCULATION) ============
    tech_total_days = 0
    dcm_total_days = 0
    global_total_days = 0
    
    for leave in LeaveRequest.objects.filter(status='approved', user__profile__team='Tech'):
        days = calculate_leave_days(leave)
        tech_total_days += days
    
    for leave in LeaveRequest.objects.filter(status='approved', user__profile__team='DCM'):
        days = calculate_leave_days(leave)
        dcm_total_days += days
    
    for leave in LeaveRequest.objects.filter(status='approved', user__profile__team='Global'):
        days = calculate_leave_days(leave)
        global_total_days += days
    
    # Calculate absence rate (assuming 21 working days per month per employee)
    tech_working_days = tech_count * 21 * 12  # 12 months
    dcm_working_days = dcm_count * 21 * 12
    global_working_days = global_count * 21 * 12
    
    tech_absence_rate = round(
        (tech_total_days / tech_working_days * 100)
    ) if tech_working_days > 0 else 0
    
    dcm_absence_rate = round(
        (dcm_total_days / dcm_working_days * 100)
    ) if dcm_working_days > 0 else 0
    
    global_absence_rate = round(
        (global_total_days / global_working_days * 100)
    ) if global_working_days > 0 else 0
    
    # ============ BALANCE SUMMARY ============
    # Calculate total available vs used days (not requests)
    # Assuming 20 days per employee per year
    total_available_leaves = total_members * 20
    total_remaining = max(0, total_available_leaves - total_days_consumed)
    
    # ============ LEAVES BY STATUS - PIE CHART DATA ============
    leaves_by_status = LeaveRequest.objects.values('status').annotate(
        count=Count('id')
    )
    
    # ============ DURATION BREAKDOWN (NEW!) ============
    full_day_count = LeaveRequest.objects.filter(
        status='approved',
        duration='full'
    ).count()
    
    first_half_count = LeaveRequest.objects.filter(
        status='approved',
        duration='first_half'
    ).count()
    
    second_half_count = LeaveRequest.objects.filter(
        status='approved',
        duration='second_half'
    ).count()
    
    duration_breakdown = [
        {'duration': 'Full Day', 'count': full_day_count},
        {'duration': 'First Half', 'count': first_half_count},
        {'duration': 'Second Half', 'count': second_half_count},
    ]
    
    # ============ PREPARE CONTEXT ============
    context = {
        # Basic counts
        'total_members': total_members,
        'tech_count': tech_count,
        'dcm_count': dcm_count,
        'global_count': global_count,
        'on_leave_today': on_leave_today,
        
        # Percentages
        'tech_percentage': tech_percentage,
        'dcm_percentage': dcm_percentage,
        
        # Leave statistics
        'total_leaves_used': total_leaves_requests,  # Number of requests
        'total_days_consumed': total_days_consumed,   # Number of days (with half-day support)
        'pending_approvals': pending_approvals,
        'rejected_leaves': rejected_leaves,
        'avg_leaves_per_emp': avg_leaves_per_emp,
        
        # Balance info
        'total_available_leaves': total_available_leaves,
        'total_remaining': total_remaining,
        
        # Team data
        'leaves_by_team': leaves_by_team,
        'pending_by_team': pending_by_team,
        'leaves_by_status': leaves_by_status,
        
        # Trends
        'monthly_trend': monthly_trend,
        'monthly_trend_days': monthly_trend_days,
        'monthly_percentages': monthly_percentages,
        'monthly_days_percentages': monthly_days_percentages,
        'months': months,
        
        # Leave types
        'leave_types': leave_type_data,
        'top_leave_takers': top_leave_takers,
        'upcoming_leaves': upcoming_leaves,
        
        # Attendance rates
        'tech_absence_rate': tech_absence_rate,
        'dcm_absence_rate': dcm_absence_rate,
        'global_absence_rate': global_absence_rate,
        
        # Duration breakdown
        'duration_breakdown': duration_breakdown,
    }
    
    return render(request, 'leaves/leave_analytics.html', context)

@login_required
@require_http_methods(["GET", "POST"])
@csrf_protect
def leave_approvals(request):
    """
    Handle leave approval requests
    GET: Display pending leave requests
    POST: Process approve/reject actions via AJAX
    """
    
    # Security check - Only managers and directors can approve
    try:
        profile = request.user.profile
        if profile.role not in ['manager', 'director']:
            if request.method == 'POST':
                return JsonResponse({
                    'success': False,
                    'message': 'You do not have permission to approve leaves'
                }, status=403)
            return redirect('dashboard')
    except Profile.DoesNotExist:
        return redirect('dashboard')
    
    # ============ HANDLE POST REQUESTS (AJAX) ============
    if request.method == 'POST':
        try:
            # Parse JSON request body
            data = json.loads(request.body)
            leave_id = data.get('leave_id')
            action = data.get('action')  # 'approve' or 'reject'
            
            # Validate inputs
            if not leave_id or not action:
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields (leave_id, action)'
                }, status=400)
            
            # Validate action
            if action not in ['approve', 'reject']:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action. Must be "approve" or "reject"'
                }, status=400)
            
            # Get the leave request
            try:
                leave = LeaveRequest.objects.get(id=leave_id)
            except LeaveRequest.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Leave request not found'
                }, status=404)
            
            # Check if already processed
            if leave.status != 'pending':
                return JsonResponse({
                    'success': False,
                    'message': f'This leave request has already been {leave.status}'
                }, status=400)
            
            # Process the action
            if action == 'approve':
                leave.status = 'approved'
                message = f'Leave request for {leave.user.get_full_name()} has been approved'
            else:  # reject
                leave.status = 'rejected'
                message = f'Leave request for {leave.user.get_full_name()} has been rejected'
            
            # Save the updated leave request
            leave.save()
            
            # Return success response
            return JsonResponse({
                'success': True,
                'message': message,
                'leave_id': leave_id,
                'new_status': leave.status
            }, status=200)
        
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid JSON request'
            }, status=400)
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            }, status=500)
    
    # ============ HANDLE GET REQUESTS (Display page) ============
    try:
        # Get pending leaves ordered by newest first
        pending_leaves = LeaveRequest.objects.filter(
            status='pending'
        ).select_related('user', 'user__profile').order_by('-created_at')
        
        context = {
            'pending_leaves': pending_leaves,
            'pending_count': pending_leaves.count(),
        }
        
        return render(request, 'leaves/approvals.html', context)
    
    except Exception as e:
        return render(request, 'leaves/approvals.html', {
            'pending_leaves': [],
            'error': str(e)
        })

@login_required
def leave_calendar(request):
    """
    Generate leave calendar with users as rows and dates as columns
    Grouped by department with holiday highlighting
    Uses correct leave types from your database
    """
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    department_filter = request.GET.get('department', 'all')
    
    if month < 1 or month > 12:
        month = today.month
    if year < 2000 or year > 2100:
        year = today.year
    
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    
    from django.contrib.auth.models import User
    from attendance.models import Holiday, Department, EmployeeProfile
    from leaves.models import LeaveRequest
    
    departments_list = Department.objects.all().order_by('name')
    employees = EmployeeProfile.objects.select_related('user', 'department').filter(is_active=True)
    
    if department_filter != 'all':
        try:
            dept = Department.objects.get(id=department_filter)
            employees = employees.filter(department=dept)
        except:
            pass
    
    employees = employees.order_by('department__name', 'user__first_name', 'user__last_name')
    
    # YOUR EXACT LEAVE TYPE MAPPING
    leave_type_mapping = {
        'Sick': {'display': 'Sick Leave', 'class': 'sick-leave', 'color': '#FFA500'},
        'Casual': {'display': 'Casual Leave', 'class': 'casual-leave', 'color': '#3498DB'},
        'Earned': {'display': 'Earned Leave', 'class': 'earned-leave', 'color': '#9B59B6'},
        'Maternity': {'display': 'Maternity Leave', 'class': 'maternity-leave', 'color': '#E91E63'},
        'WFH': {'display': 'Work From Home', 'class': 'work-from-home', 'color': '#6BCB77'},
        'Other': {'display': 'Other', 'class': 'other-leave', 'color': '#95A5A6'}
    }
    
    try:
        leaves = LeaveRequest.objects.filter(
            status='approved',
            start_date__lte=last_day,
            end_date__gte=first_day,
            user__is_active=True
        ).values('id','user_id', 'start_date', 'end_date', 'leave_type')
        leaves_list = list(leaves)
    except Exception as e:
        print(f"Error fetching leaves: {e}")
        leaves_list = []
    
    holidays_dict = {}
    try:
        holidays = Holiday.objects.filter(date__gte=first_day, date__lte=last_day)
        for holiday in holidays:
            holidays_dict[holiday.date.day] = holiday.name
    except Exception as e:
        print(f"Error fetching holidays: {e}")
    
    user_leaves = {}
    for leave in leaves_list:
        user_id = leave['user_id']
        if user_id not in user_leaves:
            user_leaves[user_id] = []
        user_leaves[user_id].append(leave)
    
    calendar_days = list(range(1, last_day.day + 1))
    department_groups = {}
    
    for emp in employees:
        user = emp.user
        dept_name = emp.department.name if emp.department else 'No Department'
        
        if dept_name not in department_groups:
            department_groups[dept_name] = []
        
        user_data = {
            'user': user,
            'employee_profile': emp,
            'leaves_by_date': {}
        }
        
        for day_num in calendar_days:
            current_date = date(year, month, day_num)
            user_data['leaves_by_date'][day_num] = None
            
            for leave in user_leaves.get(user.id, []):
                if leave['start_date'] <= current_date <= leave['end_date']:
                    leave_key = leave['leave_type']
                    
                    if leave_key in leave_type_mapping:
                        mapping = leave_type_mapping[leave_key]
                        user_data['leaves_by_date'][day_num] = {
                            'type': mapping['display'],
                            'type_class': mapping['class'],
                            'color': mapping['color'],
                            'id': leave['id']
                        }
                    else:
                        user_data['leaves_by_date'][day_num] = {
                            'type': leave_key,
                            'type_class': 'unknown-leave',
                            'color': '#CCCCCC',
                            'id': leave['id']
                        }
                    break
        
        department_groups[dept_name].append(user_data)
    
    leave_colors = {v['display']: v['color'] for v in leave_type_mapping.values()}
    
    context = {
        'department_groups': department_groups,
        'calendar_days': calendar_days,
        'holidays_dict': holidays_dict,
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month],
        'departments': departments_list,
        'selected_department': department_filter,
        'leave_colors': leave_colors,
    }
    
    return render(request, 'leaves/leave_calendar.html', context)

@login_required
@require_http_methods(["POST"])
@csrf_protect
def cancel_leave(request):
    """
    Cancel an approved leave request
    - Updates leave status to 'cancelled'
    - Returns used_leaves to available
    - Only the employee who requested or a manager/director can cancel
    """
    
    try:
        # Parse JSON request
        data = json.loads(request.body)
        leave_id = data.get('leave_id')
        
        # Validate input
        if not leave_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing leave_id'
            }, status=400)
        
        # Get the leave request
        try:
            leave = LeaveRequest.objects.get(id=leave_id)
        except LeaveRequest.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Leave request not found'
            }, status=404)
        
        # Security check - Only the employee or manager/director can cancel
        try:
            user_profile = request.user.profile
            is_manager = user_profile.role in ['manager', 'director']
        except:
            is_manager = False
        
        is_own_leave = leave.user == request.user
        
        if not (is_own_leave or is_manager):
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to cancel this leave'
            }, status=403)
        
        # Check if leave is already approved (can only cancel approved leaves)
        if leave.status != 'approved':
            return JsonResponse({
                'success': False,
                'message': f'Cannot cancel a {leave.status} leave request'
            }, status=400)
        
        # Check if leave date has already passed
        if leave.end_date < date.today():
            return JsonResponse({
                'success': False,
                'message': 'Cannot cancel a leave that has already been completed'
            }, status=400)
        
        # Calculate days that will be cancelled
        def calculate_leave_days(leave_obj):
            """Calculate total days, accounting for half days"""
            total_days = (leave_obj.end_date - leave_obj.start_date).days + 1
            if getattr(leave_obj, "duration", None) in ['first_half', 'second_half']:
                return total_days * 0.5
            return total_days
        
        days_to_cancel = (leave.end_date - leave.start_date).days + 1
        
        # Update leave status
        leave.status = 'cancelled'
        leave.save()
        
        # Prepare response
        return JsonResponse({
            'success': True,
            'message': f'Leave has been cancelled successfully. {days_to_cancel} days returned to available leaves.',
            'leave_id': leave_id,
            'days_returned': days_to_cancel,
            'new_status': 'cancelled'
        }, status=200)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON request'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'An error occurred: {str(e)}'
        }, status=500)
    
@login_required
def export_analytics_excel(request):
    """
    Export analytics dashboard data to Excel file
    """
    # Security check
    profile = request.user.profile
    if profile.role not in ['manager', 'director']:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    # Get all context data (reuse from team_analytics)
    today = timezone.now().date()
    current_month_start = today.replace(day=1)
    next_month = current_month_start + timedelta(days=32)
    current_month_end = next_month.replace(day=1) - timedelta(days=1)
    
    # ===== HELPER FUNCTION =====
    def calculate_leave_days(leave_obj):
        total_days = (leave_obj.end_date - leave_obj.start_date).days + 1
        if leave_obj.duration in ['first_half', 'second_half']:
            return total_days * 0.5
        return total_days
    
    # ===== GATHER DATA =====
    global_count = Profile.objects.filter(team='Global').count()
    tech_count = Profile.objects.filter(team='Tech').count()
    dcm_count = Profile.objects.filter(team='DCM').count()
    total_members = tech_count + dcm_count + global_count
    
    tech_percentage = round((tech_count / total_members * 100)) if total_members > 0 else 0
    dcm_percentage = round((dcm_count / total_members * 100)) if total_members > 0 else 0
    
    # Team data
    teams_data = [
        {'name': 'Tech', 'count': tech_count, 'percentage': tech_percentage},
        {'name': 'DCM', 'count': dcm_count, 'percentage': dcm_percentage},
        {'name': 'Global', 'count': global_count, 'percentage': 100 - tech_percentage - dcm_percentage},
    ]
    
    # Calculate days by team
    for team in teams_data:
        total_days = 0
        for leave in LeaveRequest.objects.filter(status='approved', user__profile__team=team['name']):
            total_days += calculate_leave_days(leave)
        team['days'] = round(total_days, 1)
    
    # Leave statistics
    total_leaves_requests = LeaveRequest.objects.filter(status='approved').count()
    pending_approvals = LeaveRequest.objects.filter(status='pending').count()
    rejected_leaves = LeaveRequest.objects.filter(status='rejected').count()
    
    # Upcoming leaves
    upcoming_leaves = LeaveRequest.objects.filter(
        status='approved',
        start_date__gte=today,
        start_date__lte=today + timedelta(days=30)
    ).select_related('user__profile').order_by('start_date')[:10]
    
    # Leave types
    leave_types = LeaveRequest.objects.filter(
        status='approved'
    ).values('leave_type').annotate(count=Count('id')).order_by('-count')
    
    # Top leave takers
    all_approved = LeaveRequest.objects.filter(status='approved').select_related('user', 'user__profile')
    employee_days = {}
    for leave in all_approved:
        emp_key = (leave.user.id, leave.user.first_name, leave.user.last_name, leave.user.profile.team)
        if emp_key not in employee_days:
            employee_days[emp_key] = 0
        employee_days[emp_key] += calculate_leave_days(leave)
    
    top_leave_takers = sorted(employee_days.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Total days consumed
    total_days_consumed = sum(calculate_leave_days(leave) for leave in LeaveRequest.objects.filter(status='approved'))
    
    # ===== CREATE EXCEL WORKBOOK =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="FFFFFF")
    subheader_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    data_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # ===== TITLE SECTION =====
    ws['A1'] = "LEAVE ANALYTICS DASHBOARD"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = center_alignment
    
    ws['A2'] = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:F2')
    
    current_row = 4
    
    # ===== SUMMARY STATISTICS =====
    ws[f'A{current_row}'] = "SUMMARY STATISTICS"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 1
    
    summary_data = [
        ['Total Members', total_members],
        ['Total Leave Requests', total_leaves_requests],
        ['Total Days Consumed', round(total_days_consumed, 1)],
        ['Pending Approvals', pending_approvals],
        ['Rejected Leaves', rejected_leaves],
    ]
    
    for data in summary_data:
        ws[f'A{current_row}'] = data[0]
        ws[f'B{current_row}'] = data[1]
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].fill = data_fill
        current_row += 1
    
    current_row += 1
    
    # ===== TEAM BREAKDOWN =====
    ws[f'A{current_row}'] = "TEAM BREAKDOWN"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 1
    
    # Headers
    headers = ['Team Name', 'Members', 'Team %', 'Days Used', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    for team in teams_data:
        ws[f'A{current_row}'] = team['name']
        ws[f'B{current_row}'] = team['count']
        ws[f'C{current_row}'] = f"{team['percentage']}%"
        ws[f'D{current_row}'] = team['days']
        ws[f'E{current_row}'] = "Active"
        
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center_alignment
            cell.fill = data_fill
        
        current_row += 1
    
    current_row += 1
    
    # ===== LEAVE TYPE DISTRIBUTION =====
    ws[f'A{current_row}'] = "LEAVE TYPE DISTRIBUTION"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1
    
    # Headers
    headers = ['Leave Type', 'Count', 'Percentage']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    total_leave_requests = sum([lt['count'] for lt in leave_types])
    for lt in leave_types:
        ws[f'A{current_row}'] = lt['leave_type']
        ws[f'B{current_row}'] = lt['count']
        ws[f'C{current_row}'] = f"{round(lt['count'] / total_leave_requests * 100, 1)}%" if total_leave_requests > 0 else "0%"
        
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center_alignment
            cell.fill = data_fill
        
        current_row += 1
    
    current_row += 1
    
    # ===== TOP LEAVE TAKERS =====
    ws[f'A{current_row}'] = "TOP LEAVE TAKERS (BY DAYS)"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1
    
    # Headers
    headers = ['Rank', 'Employee Name', 'Team', 'Total Days']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    for rank, (emp_key, days) in enumerate(top_leave_takers, 1):
        ws[f'A{current_row}'] = rank
        ws[f'B{current_row}'] = f"{emp_key[1]} {emp_key[2]}"
        ws[f'C{current_row}'] = emp_key[3]
        ws[f'D{current_row}'] = round(days, 1)
        
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center_alignment
            cell.fill = data_fill
        
        current_row += 1
    
    current_row += 1
    
    # ===== UPCOMING LEAVES =====
    ws[f'A{current_row}'] = "UPCOMING LEAVES (NEXT 30 DAYS)"
    ws[f'A{current_row}'].font = subheader_font
    ws[f'A{current_row}'].fill = subheader_fill
    ws.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    # Headers
    headers = ['Employee', 'Team', 'Start Date', 'End Date', 'Leave Type', 'Duration', 'Days']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    for leave in upcoming_leaves:
        ws[f'A{current_row}'] = f"{leave.user.first_name} {leave.user.last_name}"
        ws[f'B{current_row}'] = leave.user.profile.team
        ws[f'C{current_row}'] = leave.start_date.strftime('%d-%m-%Y')
        ws[f'D{current_row}'] = leave.end_date.strftime('%d-%m-%Y')
        ws[f'E{current_row}'] = leave.get_leave_type_display()
        ws[f'F{current_row}'] = leave.get_duration_display()
        ws[f'G{current_row}'] = leave.total_days
        
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center_alignment
            cell.fill = data_fill
        
        current_row += 1
    
    # ===== SET COLUMN WIDTHS =====
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    
    # ===== CREATE RESPONSE =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="leave_analytics_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_analytics_pdf(request):
    """
    Export analytics dashboard data to PDF file
    """
    # Security check
    profile = request.user.profile
    if profile.role not in ['manager', 'director']:
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    # Get all context data
    today = timezone.now().date()
    
    # ===== HELPER FUNCTION =====
    def calculate_leave_days(leave_obj):
        total_days = (leave_obj.end_date - leave_obj.start_date).days + 1
        if leave_obj.duration in ['first_half', 'second_half']:
            return total_days * 0.5
        return total_days
    
    # ===== GATHER DATA =====
    global_count = Profile.objects.filter(team='Global').count()
    tech_count = Profile.objects.filter(team='Tech').count()
    dcm_count = Profile.objects.filter(team='DCM').count()
    total_members = tech_count + dcm_count + global_count
    
    tech_percentage = round((tech_count / total_members * 100)) if total_members > 0 else 0
    dcm_percentage = round((dcm_count / total_members * 100)) if total_members > 0 else 0
    
    total_leaves_requests = LeaveRequest.objects.filter(status='approved').count()
    pending_approvals = LeaveRequest.objects.filter(status='pending').count()
    
    # Total days consumed
    total_days_consumed = sum(calculate_leave_days(leave) for leave in LeaveRequest.objects.filter(status='approved'))
    
    # Team data
    teams_data = []
    for team_name in ['Tech', 'DCM', 'Global']:
        team_count = Profile.objects.filter(team=team_name).count()
        if team_count > 0:
            total_days = sum(
                calculate_leave_days(leave)
                for leave in LeaveRequest.objects.filter(status='approved', user__profile__team=team_name)
            )
            teams_data.append([
                team_name,
                str(team_count),
                f"{round(team_count / total_members * 100)}%" if total_members > 0 else "0%",
                f"{round(total_days, 1)}",
            ])
    
    # Leave types
    leave_types = []
    for lt in LeaveRequest.objects.filter(status='approved').values('leave_type').annotate(count=Count('id')).order_by('-count'):
        percentage = round(lt['count'] / total_leaves_requests * 100, 1) if total_leaves_requests > 0 else 0
        leave_types.append([lt['leave_type'], str(lt['count']), f"{percentage}%"])
    
    # Top leave takers
    all_approved = LeaveRequest.objects.filter(status='approved').select_related('user', 'user__profile')
    employee_days = {}
    for leave in all_approved:
        emp_key = (leave.user.first_name, leave.user.last_name, leave.user.profile.team)
        if emp_key not in employee_days:
            employee_days[emp_key] = 0
        employee_days[emp_key] += calculate_leave_days(leave)
    
    top_leave_takers = []
    for rank, (emp_key, days) in enumerate(sorted(employee_days.items(), key=lambda x: x[1], reverse=True)[:5], 1):
        top_leave_takers.append([str(rank), f"{emp_key[0]} {emp_key[1]}", emp_key[2], f"{round(days, 1)}"])
    
    # Upcoming leaves
    upcoming_leaves = []
    for leave in LeaveRequest.objects.filter(
        status='approved',
        start_date__gte=today,
        start_date__lte=today + timedelta(days=30)
    ).select_related('user__profile').order_by('start_date')[:10]:
        upcoming_leaves.append([
            f"{leave.user.first_name} {leave.user.last_name}",
            leave.user.profile.team,
            leave.start_date.strftime('%d-%m-%Y'),
            leave.end_date.strftime('%d-%m-%Y'),
            leave.get_leave_type_display(),
            leave.get_duration_display(),
            str(leave.total_days),
        ])
    
    # ===== CREATE PDF =====
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1565C0'),
        spaceAfter=12,
        alignment=1,  # Center
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=8,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    # ===== TITLE =====
    story.append(Paragraph("LEAVE ANALYTICS DASHBOARD", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # ===== SUMMARY STATISTICS =====
    story.append(Paragraph("SUMMARY STATISTICS", heading_style))
    
    summary_table_data = [
        ['Total Members', str(total_members)],
        ['Total Leave Requests', str(total_leaves_requests)],
        ['Total Days Consumed', f"{round(total_days_consumed, 1)}"],
        ['Pending Approvals', str(pending_approvals)],
    ]
    
    summary_table = Table(summary_table_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # ===== TEAM BREAKDOWN =====
    story.append(Paragraph("TEAM BREAKDOWN", heading_style))
    
    team_table_data = [['Team Name', 'Members', 'Team %', 'Days Used']]
    team_table_data.extend(teams_data)
    
    team_table = Table(team_table_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1.5*inch])
    team_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
    ]))
    story.append(team_table)
    story.append(Spacer(1, 0.3*inch))
    
    # ===== LEAVE TYPE DISTRIBUTION =====
    story.append(Paragraph("LEAVE TYPE DISTRIBUTION", heading_style))
    
    leave_type_table_data = [['Leave Type', 'Count', 'Percentage']]
    leave_type_table_data.extend(leave_types)
    
    leave_type_table = Table(leave_type_table_data, colWidths=[2*inch, 1*inch, 1*inch])
    leave_type_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
    ]))
    story.append(leave_type_table)
    story.append(Spacer(1, 0.3*inch))
    
    # ===== TOP LEAVE TAKERS =====
    story.append(Paragraph("TOP LEAVE TAKERS (BY DAYS)", heading_style))
    
    top_takers_table_data = [['Rank', 'Employee Name', 'Team', 'Days']]
    top_takers_table_data.extend(top_leave_takers)
    
    top_takers_table = Table(top_takers_table_data, colWidths=[0.8*inch, 2*inch, 1.2*inch, 1*inch])
    top_takers_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
    ]))
    story.append(top_takers_table)
    
    if upcoming_leaves:
        story.append(PageBreak())
        story.append(Paragraph("UPCOMING LEAVES (NEXT 30 DAYS)", heading_style))
        
        upcoming_table_data = [['Employee', 'Team', 'Start Date', 'End Date', 'Leave Type', 'Duration', 'Days']]
        upcoming_table_data.extend(upcoming_leaves)
        
        upcoming_table = Table(upcoming_table_data, colWidths=[1.2*inch, 0.8*inch, 1*inch, 1*inch, 1.2*inch, 1*inch, 0.8*inch])
        upcoming_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
        ]))
        story.append(upcoming_table)
    
    # ===== BUILD PDF =====
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="leave_analytics_{datetime.now().strftime("%d_%m_%Y")}.pdf"'
    response.write(buffer.getvalue())
    
    return response